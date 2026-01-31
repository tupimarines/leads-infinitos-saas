import datetime

from dataclasses import dataclass, asdict, field
import pandas as pd
import argparse
import os
import sys
import re
from typing import List, Dict
import random

from apify_client import ApifyClient

@dataclass
class Business:
    """holds business data"""
    name: str = None
    address: str = None
    domain: str = None
    website: str = None
    phone_number: str = None
    whatsapp_link: str = None
    category: str = None
    location: str = None
    reviews_count: int = None
    reviews_average: float = None
    latitude: float = None
    longitude: float = None
    
    def __hash__(self):
        """Make Business hashable for duplicate detection.
        Consider businesses different if:
        - Name is different, OR
        - Same name but different non-empty contact info (domain/website/phone)
        """
        hash_fields = [self.name]
        if self.domain:
            hash_fields.append(f"domain:{self.domain}")
        if self.website:
            hash_fields.append(f"website:{self.website}")
        if self.phone_number:
            hash_fields.append(f"phone:{self.phone_number}")
        
        return hash(tuple(hash_fields))

@dataclass
class BusinessList:
    """Holds list of Business objects and saves to both Excel and CSV.

    The output directory is based on `save_base_dir` and the current date.
    For example: <save_base_dir>/<YYYY-MM-DD>/...
    """
    business_list: list[Business] = field(default_factory=list)
    _seen_businesses: set = field(default_factory=set, init=False)
    save_base_dir: str = 'GMaps Data'
    today: str = field(default_factory=lambda: datetime.datetime.now().strftime("%Y-%m-%d"))
    save_at: str = field(init=False)

    def __post_init__(self):
        self.save_at = os.path.join(self.save_base_dir, self.today)
        os.makedirs(self.save_at, exist_ok=True)

    def add_business(self, business: Business):
        """Add a business to the list if it's not a duplicate based on key attributes"""
        business_hash = hash(business)
        if business_hash not in self._seen_businesses:
            self.business_list.append(business)
            self._seen_businesses.add(business_hash)
    
    def dataframe(self, add_status_column: bool = False):
        """transform business_list to pandas dataframe

        Args:
            add_status_column: Se True, adiciona coluna 'status' com valor 1

        Returns: pandas dataframe
        """
        df = pd.json_normalize(
            (asdict(business) for business in self.business_list), sep="_"
        )
        
        # Adicionar coluna status se solicitado
        if add_status_column:
            df['status'] = 1
        
        return df

    def save_to_excel(self, filename):
        """saves pandas dataframe to excel (xlsx) file

        Args:
            filename (str): filename
        """
        try:
            df = self.dataframe()
            out_path = f"{self.save_at}/{filename}.xlsx"
            # Write with openpyxl engine so we can post-process hyperlinks
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False)
                try:
                    from openpyxl.utils import get_column_letter
                    ws = writer.book.active
                    if "whatsapp_link" in df.columns:
                        col_idx = df.columns.get_loc("whatsapp_link") + 1  # 1-based
                        col_letter = get_column_letter(col_idx)
                        for row_idx in range(2, len(df) + 1):  # skip header
                            cell = ws[f"{col_letter}{row_idx}"]
                            link = cell.value
                            if link:
                                cell.hyperlink = link
                                cell.style = "Hyperlink"
                except Exception:
                    # If anything goes wrong, keep the plain values without hyperlinks
                    pass
        except ImportError:
            print("openpyxl not installed; skipping Excel export and continuing with CSV...")
        except Exception as e:
            print(f"Failed to write Excel: {e}; continuing with CSV...")

    def save_to_csv(self, filename):
        """saves pandas dataframe to csv file

        Args:
            filename (str): filename
        """
        self.dataframe().to_csv(f"{self.save_at}/{filename}.csv", index=False)

    def save_to_excel_with_status(self, filename):
        """
        Salva dataframe com coluna status adicionada

        Args:
            filename (str): filename
        """
        try:
            df = self.dataframe(add_status_column=True)
            out_path = f"{self.save_at}/{filename}.xlsx"
            # Write with openpyxl engine so we can post-process hyperlinks
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False)
                try:
                    from openpyxl.utils import get_column_letter
                    ws = writer.book.active
                    if "whatsapp_link" in df.columns:
                        col_idx = df.columns.get_loc("whatsapp_link") + 1  # 1-based
                        col_letter = get_column_letter(col_idx)
                        for row_idx in range(2, len(df) + 1):  # skip header
                            cell = ws[f"{col_letter}{row_idx}"]
                            link = cell.value
                            if link:
                                cell.hyperlink = link
                                cell.style = "Hyperlink"
                except Exception:
                    # If anything goes wrong, keep the plain values without hyperlinks
                    pass
        except ImportError:
            print("openpyxl not installed; skipping Excel export and continuing with CSV...")
        except Exception as e:
            print(f"Failed to write Excel: {e}; continuing with CSV...")

    def save_to_csv_with_status(self, filename):
        """saves pandas dataframe to csv file with status column

        Args:
            filename (str): filename
        """
        self.dataframe(add_status_column=True).to_csv(f"{self.save_at}/{filename}.csv", index=False)

def concatenate_business_lists(business_lists: List[BusinessList]) -> BusinessList:
    """
    Concatena múltiplas BusinessList em uma única, com deduplicação automática
    """
    if not business_lists:
        return BusinessList()
    
    # Usar a primeira BusinessList como base
    result = business_lists[0]
    
    # Adicionar businesses das outras listas
    for business_list in business_lists[1:]:
        for business in business_list.business_list:
            result.add_business(business)  # Deduplicação automática
    
    return result






def run_apify_scraper(search_list, total, save_base_dir, progress_callback=None) -> List[Dict[str, str]]:
    """
    Executes Google Maps scraping using Apify Actor (compass/crawler-google-places).
    Maps results to our Business format and saves to Excel/CSV.
    """
    apify_token = os.getenv("APIFY_TOKEN")
    if not apify_token:
        raise ValueError("APIFY_TOKEN not found in environment variables")
        
    client = ApifyClient(apify_token)
    
    results = []
    
    for i, search_item in enumerate(search_list):
        # Determine if input is simple string or structured dict
        if isinstance(search_item, dict):
            keyword = search_item.get('keyword')
            location = search_item.get('location')
            search_term_display = f"{keyword} in {location}"
        else:
            keyword = search_item
            location = None
            search_term_display = search_item

        print(f"Executing Apify Actor for: {search_term_display} ({i+1}/{len(search_list)})")
        
        if progress_callback:
            try:
                progress = int((i / len(search_list)) * 100)
                progress_callback(progress, search_term_display)
            except:
                pass

        # Prepare input for Apify
        run_input = {
            "searchStringsArray": [keyword],
            "maxCrawledPlacesPerSearch": total,
            "language": "pt-BR", 
            "scrapeSocialMediaProfiles": {
                "facebooks": False,
                "instagrams": False,
                "youtubes": False,
                "tiktoks": False,
                "twitters": False,
            },
            "maximumLeadsEnrichmentRecords": 0,
            "maxImages": 0,
            "countryCode": "br",
        }
        
        # Add locationQuery if present (stricter geographical search)
        if location:
            run_input["locationQuery"] = location

        try:
            # Run the Actor and wait for it to finish
            print(f"Calling Apify Actor with input: keyword='{keyword}', location='{location}'")
            run = client.actor("compass/crawler-google-places").call(run_input=run_input)
            
            print(f"Apify Run Finished. Dataset ID: {run['defaultDatasetId']}")
            
            dataset_items = client.dataset(run["defaultDatasetId"]).list_items().items
            print(f"Retrieved {len(dataset_items)} items from Apify dataset")
            
            # Map to Business Objects
            b_list = BusinessList(save_base_dir=save_base_dir)
            
            for item in dataset_items:
                # Map fields
                # Apify fields are diverse (title, address, phoneUnformatted, website, etc)
                
                try:
                    category_str = item.get("categoryName") or (item.get("categories")[0] if item.get("categories") else None)
                    
                    addr = item.get("address")
                    
                    b = Business(
                        name=item.get("title"),
                        address=addr,
                        domain=None, # TBD if needed
                        website=item.get("website"),
                        phone_number=item.get("phoneUnformatted") or item.get("phone"),
                        whatsapp_link=None, # Apify doesn't give WA link directly usually
                        category=category_str,
                        location=search_term_display, # Or derive from address
                        reviews_count=item.get("reviewsCount"),
                        reviews_average=item.get("totalScore"),
                        latitude=item.get("location", {}).get("lat"),
                        longitude=item.get("location", {}).get("lng")
                    )
                    b_list.add_business(b)
                except Exception as map_err:
                    print(f"Error mapping item: {map_err}")

            # Save results for this search term WITH STATUS COLUMN
            safe_filename = re.sub(r'[^a-zA-Z0-9]', '_', search_term_display)
            b_list.save_to_excel_with_status(safe_filename)
            b_list.save_to_csv_with_status(safe_filename)
            
            results.append({
                "search": search_term_display,
                "csv_path": f"{b_list.save_at}/{safe_filename}.csv",
                "xlsx_path": f"{b_list.save_at}/{safe_filename}.xlsx"
            })
            
        except Exception as e:
            print(f"Apify Actor Failed for {search_term}: {e}")
            # Continue to next search term
    
    return results




def run_scraper_with_progress(
    search_list: List[str],
    total: int,
    headless: bool = True,
    save_base_dir: str | None = None,
    concatenate_results: bool = False,
    progress_callback = None,
) -> List[Dict[str, str]]:
    """
    Wrapper function called by worker to execute Apify Scraper.
    Arguments like 'headless' are ignored as Apify runs in the cloud.
    """
    return run_apify_scraper(search_list, total, save_base_dir, progress_callback)




def main():
    # read search from arguments
    parser = argparse.ArgumentParser()
    parser.add_argument("-s", "--search", type=str)
    parser.add_argument("-t", "--total", type=int)
    args = parser.parse_args()
    
    if args.search:
        search_list = [args.search]
    
    if args.total:
        total = args.total
    else:
        total = 1_000_000

    if not args.search:
        search_list = []
        input_file_name = 'input.txt'
        input_file_path = os.path.join(os.getcwd(), input_file_name)
        if os.path.exists(input_file_path):
            with open(input_file_path, 'r') as file:
                search_list = file.readlines()
                
        if len(search_list) == 0:
            print('Error occured: You must either pass the -s search argument, or add searches to input.txt')
            sys.exit()
    
    # CLI uses headful browser for visibility
    run_scraper(search_list=search_list, total=total, headless=False)

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f'Failed err: {e}')
